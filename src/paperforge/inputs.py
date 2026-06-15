"""Read DOIs (and light metadata) from .xlsx / .csv / .txt inputs.

Strategy per tabular file:
  1. Look at the header row for a column whose name contains "doi".
     If found, read that column (plus author/year/title columns when present).
  2. If there is no DOI header, regex-scan every cell for DOI-looking strings.

Plain-text files are scanned line by line. Heterogeneous spreadsheets (any
sheet, any column order) are handled because matching is by header *name*,
with a regex fallback when names don't help.
"""
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from .utils import clean_year, extract_dois


@dataclass
class PaperRecord:
    doi: str
    author: str = ""
    year: str = ""
    title: str = ""
    origin: str = ""   # e.g. "refs.xlsx:Sheet1:row12" — for traceability


_DOI_KEYS = ("doi",)
_AUTHOR_KEYS = ("author", "creator")
_YEAR_KEYS = ("year", "published", "publication date", "pub date", "date")
_TITLE_KEYS = ("title",)


def _match_columns(header: list[str]) -> dict[str, int]:
    """Map field -> column index by header-name matching. Empty if no DOI column."""
    norm = [(h or "").strip().lower() for h in header]

    def find(keys: tuple[str, ...]):
        for i, h in enumerate(norm):       # exact match wins
            if h in keys:
                return i
        for i, h in enumerate(norm):       # then substring
            if h and any(k in h for k in keys):
                return i
        return None

    di = find(_DOI_KEYS)
    if di is None:
        return {}
    cols = {"doi": di}
    for field_name, keys in (("author", _AUTHOR_KEYS),
                             ("year", _YEAR_KEYS),
                             ("title", _TITLE_KEYS)):
        idx = find(keys)
        if idx is not None:
            cols[field_name] = idx
    return cols


def _cell(row, idx) -> str:
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def _rows_to_records(rows: list[list], origin_prefix: str) -> list[PaperRecord]:
    rows = [r for r in rows]
    if not rows:
        return []

    header = [("" if c is None else str(c)) for c in rows[0]]
    cols = _match_columns(header)
    records: list[PaperRecord] = []

    if cols:
        di = cols["doi"]
        for n, row in enumerate(rows[1:], start=2):
            # Validate via the DOI regex; this also unwraps URL/`doi:` cells and
            # rejects stray non-DOI values that happen to sit in the column.
            found = extract_dois(_cell(row, di))
            if not found:
                continue
            doi = found[0]
            year_raw = _cell(row, cols.get("year"))
            records.append(PaperRecord(
                doi=doi,
                author=_cell(row, cols.get("author")),
                year=clean_year(year_raw) or year_raw,
                title=_cell(row, cols.get("title")),
                origin=f"{origin_prefix}:row{n}",
            ))
    else:
        # No DOI header anywhere: regex-scan every cell of every row.
        for n, row in enumerate(rows, start=1):
            text = " ".join("" if c is None else str(c) for c in row)
            for doi in extract_dois(text):
                records.append(PaperRecord(doi=doi, origin=f"{origin_prefix}:row{n}"))

    return records


def _load_csv(path: Path) -> list[PaperRecord]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel
        rows = [list(r) for r in csv.reader(f, dialect)]
    return _rows_to_records(rows, path.name)


def _load_xlsx(path: Path) -> list[PaperRecord]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        records: list[PaperRecord] = []
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            records.extend(_rows_to_records(rows, f"{path.name}:{ws.title}"))
        return records
    finally:
        wb.close()


def _load_txt(path: Path) -> list[PaperRecord]:
    records: list[PaperRecord] = []
    with open(path, encoding="utf-8", errors="replace") as f:
        for n, line in enumerate(f, start=1):
            for doi in extract_dois(line):
                records.append(PaperRecord(doi=doi, origin=f"{path.name}:line{n}"))
    return records


def load_records(path) -> list[PaperRecord]:
    """Load PaperRecords from a single file, dispatching on extension."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _load_xlsx(path)
    if suffix in (".csv", ".tsv"):
        return _load_csv(path)
    if suffix == ".xls":
        raise ValueError(f"{path.name}: legacy .xls is unsupported — re-save as .xlsx")
    # .txt, no extension, or anything else: treat as text and scan for DOIs.
    return _load_txt(path)


def load_many(paths: Iterable) -> list[PaperRecord]:
    """Load and concatenate records from several files (dedup happens later)."""
    out: list[PaperRecord] = []
    for p in paths:
        out.extend(load_records(p))
    return out
